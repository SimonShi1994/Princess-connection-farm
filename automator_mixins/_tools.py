import datetime
import os
import random
import time
from typing import Optional

import cv2
import numpy as np
import openpyxl
import pandas as pd
import pathlib
from scenes.root.juese import get_name_from_plate_path

from automator_mixins._base import DEBUG_RECORD
from core.MoveRecord import movevar
from core.constant import MAIN_BTN, PCRelement, ZHUCAIDAN_BTN, JUESE_BTN, JUQING_BTN, p
from core.constant import USER_DEFAULT_DICT as UDD
from core.cv import UIMatcher
from core.log_handler import pcr_log
from core.pcr_config import debug, fast_screencut, lockimg_timeout, use_pcrocr_to_process_basic_text
from core.safe_u2 import timeout
from core.tkutils import TimeoutMsgBox
from core.usercentre import get_all_group
from core.utils import make_it_as_number_as_possible, make_it_as_zhuangbei_as_possible, get_time_str, checkNameValid
from ._base import BaseMixin
from scenes.root.juese import CharMenu


class ToolsMixin(BaseMixin):
    """
    工具类插片
    包含一些辅助功能和辅助类脚本
    还有很多常用函数，比如回首页
    """

    @DEBUG_RECORD
    def lock_home(self):
        """
        锁定首页
        要求场景：存在“我的主页”按钮
        逻辑：不断点击我的主页，直到右下角出现“礼物”
        """
        self.clear_all_initFC()
        last = time.time()
        while True:
            time.sleep(1)
            sc = self.getscreen()
            if self.is_exists(MAIN_BTN["xiazai"], screen=sc):
                self.click(MAIN_BTN["xiazai"])
            num_of_white, _, x, y = UIMatcher.find_gaoliang(sc)
            if num_of_white < 77000:
                self.chulijiaocheng(None)  # 增加对教程的处理功能
                last = time.time()
            try:
                r_list = self.img_where_all(img=MAIN_BTN["guanbi"], screen=sc, threshold=0.90)
                if self.lock_no_img(img=MAIN_BTN["guanbi"], elseclick=(int(r_list[0]), int(r_list[1])),
                                    side_check=self.juqing_kkr):
                    time.sleep(2)
                    continue
            except:
                pass
            if self.is_exists(MAIN_BTN["liwu"], screen=sc):
                return
            # if self.is_exists(MAIN_BTN["guanbi"], screen=sc):
            #     self.click(MAIN_BTN["guanbi"])
            self.fclick(MAIN_BTN["zhuye"])
            # 防卡公告
            self.fclick(1, 1)
            time.sleep(1.5)
            if time.time() - last > lockimg_timeout:
                raise Exception("lock_home时出错：超时！")

    def get_zhuye(self):
        """
        锁定主页后
        返回主页Scene
        """
        from scenes.root.wodezhuye import WoDeZhuYe
        self.lock_home()
        return WoDeZhuYe(self).enter(timeout=300)

    @timeout(300, "init_home执行超时：超过5分钟")
    @DEBUG_RECORD
    def init_home(self):
        # 2020-07-31 TheAutumnOfRice: 检查完毕
        cnt = 0
        while True:
            time.sleep(2)
            screen_shot_ = self.getscreen()
            if self.is_exists(MAIN_BTN["liwu"], screen=screen_shot_):
                break
            if self.is_exists(MAIN_BTN["xiazai"], screen=screen_shot_):
                self.click(MAIN_BTN["xiazai"])
            if self.is_exists(MAIN_BTN["tiaoguo"], screen=screen_shot_):
                self.click(893, 39, post_delay=0.5)  # 跳过
                continue
            if self.is_exists(MAIN_BTN["xzcw"], screen=screen_shot_):
                raise Exception("下载错误")
            if self.is_exists(MAIN_BTN["jingsaikaishi"], screen=screen_shot_):
                self.click(786, 308, post_delay=0.2)  # 选角色
                self.click(842, 491)  # 开始
                continue
            num_of_white, _, x, y = UIMatcher.find_gaoliang(screen_shot_)
            if num_of_white < 77000:
                cnt += 1
                time.sleep(1.5)  # 防止黑屏错误识别
                if cnt >= 2:
                    self.chulijiaocheng(None)  # 增加对教程的处理功能
            try:
                r_list = self.img_where_all(img=MAIN_BTN["guanbi"], screen=screen_shot_)
                if self.lock_no_img(img=MAIN_BTN["guanbi"], elseclick=(int(r_list[0]), int(r_list[1])),
                                    side_check=self.juqing_kkr):
                    time.sleep(3)
                    continue
            except:
                pass
            # if self.is_exists(MAIN_BTN["guanbi"], screen=screen_shot_):
            #     self.click(MAIN_BTN["guanbi"])
            #     continue
            # num_of_white, _, x, y = UIMatcher.find_gaoliang(screen_shot_)
            # print(num_of_white)
            # if num_of_white < 77000:
            #     break

            # 跳过特别庆典
            self.click(1, 1, post_delay=0.5)
            self.click(330, 270, post_delay=1)
            # 跳过抽签（备用）
            self.d.touch.down(370, 330).sleep(0.1).move(50, 50).sleep(0.2).up(370, 450)

        self.lock_home()
        time.sleep(0.5)
        # 这里防一波第二天可可萝跳脸教程
        screen_shot_ = self.getscreen()
        num_of_white, _, _, _ = UIMatcher.find_gaoliang(screen_shot_)
        if num_of_white < 50000:
            self.lock_img('img/renwu_1.bmp', elseclick=[(837, 433)], elsedelay=1)
            self.lock_home()
            return
        if UIMatcher.img_where(screen_shot_, 'img/kekeluo.bmp'):
            self.lock_img('img/renwu_1.bmp', elseclick=[(837, 433)], elsedelay=1)
            self.lock_home()
        time.sleep(1)
        self.lock_home()  # 追加检测

    def setting(self):
        self.lock_home()
        self.click_btn(MAIN_BTN["zhucaidan"], until_appear=MAIN_BTN["setting_pic"])
        self.click_btn(MAIN_BTN["setting_pic"])
        self.click(769, 87)
        time.sleep(1)
        self.click(710, 226)
        time.sleep(0.5)
        self.click(710, 349)
        time.sleep(0.5)
        self.click(479, 479)
        time.sleep(1)
        self.click(95, 516)
        self.lock_home()

    def zanting(self):
        TimeoutMsgBox("暂停", desc=f"{self.address}暂停中\n账号：{self.account}", join=True, geo="200x60")

    def maizhuangbei(self, day_interval):
        """
        卖掉数量前三的装备，（如果超过1000）
        适合小号
        :param day_interval: 日期间隔：每过day_interval天进行一次卖出
        """

        def get_last_record():
            ts = self.AR.get("time_status", UDD["time_status"])
            return ts["maizhuangbei"]

        def set_last_record():
            ts = self.AR.get("time_status", UDD["time_status"])
            ts["maizhuangbei"] = time.time()
            self.AR.set("time_status", ts)

        tm = get_last_record()
        diff = time.time() - tm
        if diff < day_interval * 3600 * 24:
            self.log.write_log("info", f"离下次卖装备还有{day_interval - int(diff / 3600 / 24)}天，跳过。")
            return
        self.lock_home()
        self.lock_img(ZHUCAIDAN_BTN["bangzhu"], elseclick=[(871, 513)])  # 锁定帮助
        self.click_btn(ZHUCAIDAN_BTN["daoju"], until_appear=ZHUCAIDAN_BTN["daojuyilan"])
        self.click_btn(ZHUCAIDAN_BTN["zhuangbei"], until_appear=ZHUCAIDAN_BTN["chushou"])
        if not self.is_exists(ZHUCAIDAN_BTN["chiyoushu"]):
            self.click(723, 32, post_delay=3)
            self.click(285, 228, post_delay=1)
            self.click(587, 377, post_delay=3)
        self.click_btn(ZHUCAIDAN_BTN["jiangxu"], until_appear=ZHUCAIDAN_BTN["jiangxu"])
        for _ in range(3):
            self.click_btn(ZHUCAIDAN_BTN["chushou"], until_appear=ZHUCAIDAN_BTN["chushouqueren"])
            self.click(645, 315, post_delay=2)  # max
            th_at = (518, 267, 530, 282)  # 千位
            img = self.getscreen()
            cut_img = UIMatcher.img_cut(img, th_at)
            if debug:
                self.log.write_log('debug', "VAR:" + cut_img.var())
            if cut_img.var() > 1000:
                # 有千位，卖
                self.click_btn(ZHUCAIDAN_BTN["chushou2"], until_appear=ZHUCAIDAN_BTN["chushouwanbi"])
                for _ in range(5):
                    self.click(1, 1)
            else:
                break
        set_last_record()
        self.lock_home()

    def get_base_info(self, base_info=False, introduction_info=False, props_info=False, char_info=False, out_xls=False,
                      s_sent=False, acc_nature=0):
        """
        账号基本信息获取
        By:CyiceK
        有bug请反馈
        :param acc_nature: 小/大号
        :param s_sent: 是否发送到Server酱
        :param out_xls: 是否输出为Excel表格
        :param base_info: 是否读取主页面的基本信息
        :param introduction_info: 是否读取介绍的基本信息
        :param props_info: 是否读取道具的基本信息-扫荡券
        :param char_info: 是否记录三星角色
        :return: acc_info_dict
        """
        # 笨方法转化时间戳"%Y-%m-%d-%H-%M-%S"
        date_start = datetime.datetime(1899, 12, 30)
        date_now = datetime.datetime.now()
        delta = date_now - date_start
        # 时间戳
        date_1900 = float(delta.days) + (float(delta.seconds) / 86400)
        # 日期
        date = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")

        acc_info_dict = {
            "dengji": 'None',
            "jianjie_name": 'None',
            "tili": 'None',
            "mana": 'None',
            "baoshi": 'None',
            "jianjie_zhanli": 'None',
            "jianjie_hanghui": 'None',
            "jianjie_id": 'None',
            "zhanghao": self.account,
            "group": ','.join(get_all_group(self.account)),
            "charlist": 'None',
            "saodangquan": 'None',
            "date": date,
        }
        acc_info_list = []
        self.lock_home()
        if base_info:
            time.sleep(2)
            self.lock_home()
            screen_shot = self.getscreen()
            # 体力 包括/
            A, B = self.ocr_A_B(243, 6, 305, 22, screen_shot=screen_shot)
            acc_info_dict["tili"] = f"{A}/{B}"
            # 等级 make_it_as_number_as_possible
            acc_info_dict["dengji"] = self.ocr_int(29, 43, 60, 67, screen_shot=screen_shot)
            # mana
            if use_pcrocr_to_process_basic_text:
                acc_info_dict["mana"] = make_it_as_number_as_possible(
                    self.ocr_center(107, 54, 177, 76, screen_shot=screen_shot, custom_ocr="pcr", allowstr="0123456789,") \
                        .replace(',', ''))
            else:
                acc_info_dict["mana"] = make_it_as_number_as_possible(
                    self.ocr_center(107, 54, 177, 76, screen_shot=screen_shot, size=2.0) \
                        .replace(',', '').replace('.', ''))
            # 宝石
            if use_pcrocr_to_process_basic_text:
                acc_info_dict["baoshi"] = make_it_as_number_as_possible(
                    self.ocr_center(258, 52, 306, 72, screen_shot=screen_shot, custom_ocr="pcr", allowstr="0123456789,") \
                        .replace(',', ''))
            else:
                acc_info_dict["baoshi"] = make_it_as_number_as_possible(
                    self.ocr_center(258, 52, 306, 72, screen_shot=screen_shot, size=2.0) \
                        .replace(',', '').replace('.', ''))
        if introduction_info:
            self.lock_img(ZHUCAIDAN_BTN["bangzhu"], elseclick=[(871, 513)])  # 锁定帮助
            # 去简介
            self.lock_no_img(ZHUCAIDAN_BTN["jianjie"], elseclick=[(382, 230)])
            self.lock_img(ZHUCAIDAN_BTN["jianjie_L"], elseclick=[(382, 230)])  # 锁定简介
            screen_shot = self.getscreen()
            acc_info_dict["jianjie_name"] = self.ocr_center(608, 151, 879, 178, screen_shot=screen_shot, size=2.0)
            acc_info_dict["dengji"] = self.ocr_int(702, 184, 785, 205, screen_shot=screen_shot)
            acc_info_dict["jianjie_zhanli"] = self.ocr_int(702, 214, 786, 235, screen_shot=screen_shot)
            acc_info_dict["jianjie_hanghui"] = self.ocr_center(703, 243, 918, 266, screen_shot=screen_shot,
                                                               size=2.0)
            acc_info_dict["jianjie_id"] = str(self.ocr_int(598, 415, 768, 435, screen_shot=screen_shot))
        if props_info:
            self.lock_img(ZHUCAIDAN_BTN["bangzhu"], elseclick=[(871, 513)])  # 锁定帮助
            # 去道具
            self.lock_no_img(ZHUCAIDAN_BTN["daoju"], elseclick=[(536, 159)])
            self.lock_img(ZHUCAIDAN_BTN["daojuyilan"], elseclick=[(536, 159)])  # 锁定道具一览
            screen_shot = self.getscreen()
            self.click_img(screen=screen_shot, img="img/zhucaidan/saodangquan.bmp")
            time.sleep(2)
            screen_shot = self.getscreen()
            acc_info_dict["saodangquan"] = self.get_daoju_number(screen_shot, must_int=False)
        if char_info:
            self.lock_home()
            self.click_btn(MAIN_BTN["juese"], until_appear=JUESE_BTN["duiwu"])
            at1 = (38, 79, 314, 206)
            at2 = (334, 79, 610, 206)
            at3 = (628, 79, 911, 206)
            at4 = (38, 228, 314, 350)
            at5 = (334, 228, 610, 350)
            at6 = (633, 228, 911, 350)
            at_list = [at1, at2, at3, at4, at5, at6]
            charlist = []
            P = pathlib.Path("img/juese/plate/")
            cm = CharMenu(self)
            cm.sort_by(cat="star")
            while True:
                sc = self.getscreen()
                for p in P.iterdir():
                    p = str(p).replace('\\', '/')
                    if p[-4:] == ".bmp":
                        for area in at_list:
                            if self.is_exists(img=p, at=area, screen=sc):
                                name = get_name_from_plate_path(p)
                                charlist.append(name)

                if cm.check_buttom() is True:
                    break
                else:
                    cm.dragdown()
                    if self.is_exists(JUESE_BTN["weijiesuo_w"], at=(21, 144, 167, 463)):
                        break
                    continue

            out = ','.join(charlist)
            acc_info_dict["charlist"] = out
        acc_info_list.append(acc_info_dict)
        self.lock_home()
        # 表格数据整理和转换
        if out_xls:
            # 将字典列表转换为DataFrame
            pf = pd.DataFrame(list(acc_info_list))
            # 指定字段顺序
            order = ['dengji', 'jianjie_name', 'tili', 'mana', 'baoshi', 'jianjie_zhanli',
                     'jianjie_hanghui', 'jianjie_id', 'zhanghao', 'group', 'charlist', 'saodangquan', 'date']
            pf = pf[order]
            # 将列名替换为中文
            columns_map = {
                'dengji': '等级',
                'jianjie_name': '名字',
                'tili': '体力',
                'mana': '玛娜数量',
                'baoshi': '宝石数量',
                'jianjie_zhanli': '全角色战力',
                'jianjie_hanghui': '所属行会',
                'jianjie_id': '玩家ID',
                'zhanghao': '账号',
                'group': '所在组',
                'charlist': '持有角色',
                'saodangquan': '所拥有的扫荡券',
                'date': '录入日期',
            }
            pf.rename(columns=columns_map, inplace=True)

            if acc_nature == 0:
                # 小号/农场号输出格式
                xls_path = 'xls/%s-pcr_farm_info.xlsx' % self.today_date
            elif acc_nature == 1:
                # 大号统一文件格式
                xls_path = 'xls/pcr_farm_info.xlsx'
            else:
                # 乱输入就这样的格式
                xls_path = 'xls/%s-pcr_farm_info.xlsx' % self.today_date

            # 将空的单元格替换为空字符
            pf.fillna('', inplace=True)
            # 判断文件是否存在
            if not os.path.exists(xls_path):
                # 输出
                # 指定生成的Excel表格名称
                file_path = pd.ExcelWriter(xls_path, engine='openpyxl')
                pf.to_excel(file_path, engine='openpyxl', encoding='utf-8', index=False)
                # 保存表格
                file_path.save()
                return acc_info_dict
            # 多进程怎么加锁QAQ

            # 保存表格
            index = len(list(acc_info_list))  # 获取需要写入数据的行数
            workbook = openpyxl.load_workbook(xls_path)  # 打开表格
            sheets = workbook.sheetnames  # 获取表格中的所有表格
            worksheet = workbook[sheets[0]]  # 获取表格中所有表格中的的第一个表格
            rows_old = worksheet.max_row  # 获取表格中已存在的数据的行数
            for i in range(0, index):
                for j in range(0, len(list(acc_info_list)[i])):
                    # 追加写入数据，注意是从i+rows_old行开始写入
                    worksheet.cell(row=i + 1 + rows_old, column=1 + j).value = list(acc_info_dict.values())[j]
            workbook.save(xls_path)  # 保存表格

        return acc_info_dict

    def get_tili(self):
        # 利用baiduOCR获取当前体力值（要保证当前界面有‘主菜单’选项）
        # API key存放在baiduocr.txt中
        # 格式：apiKey secretKey（中间以一个\t作为分隔符）
        # 返回值：一个int类型整数；如果读取失败返回-1

        self.click(871, 513)  # 主菜单
        while True:  # 锁定帮助
            screen_shot_ = self.getscreen()
            if UIMatcher.img_where(screen_shot_, 'img/zhucaidan/bangzhu.bmp'):
                break
        # cv2.imwrite('all.png',screen_shot_)
        # part = screen_shot_[526:649, 494:524]
        ret = self.baidu_ocr(494, 526, 524, 649, 1)  # 获取体力区域的ocr结果
        if ret == -1:
            self.log.write_log('error', '体力识别失败！')
            return -1
        else:
            return int(ret['words_result'][1]['words'].split('/')[0])

    def rename(self, name, auto_id):  # 重命名
        # 2021/1/4 CyiceK对代码进行了维护
        name = name.split(' ')
        name_len = len(name)
        if auto_id:
            name = name[random.randint(0, name_len - 1)] + str(random.randint(0, 1000))
        else:
            name = name[random.randint(0, name_len - 1)]
        self.click(871, 513)  # 主菜单
        self.lock_img('img/zhucaidan/bangzhu.bmp', ifclick=[(370, 250)])  # 锁定帮助 点击简介
        self.lock_img('img/bianji.bmp', ifclick=[(900, 155)])  # 锁定 点击铅笔修改按钮
        self.lock_img('img/biangeng.bmp', ifclick=[(480, 270)])  # 锁定 玩家名 点击游戏渲染编辑框
        time.sleep(1)
        self.click(290, 425)  # 点击编辑框
        self.d.clear_text()
        self.d.send_keys(name)
        self.click(880, 425)  # 点击确定
        time.sleep(0.5)
        self.click(590, 370)  # 变更按钮
        time.sleep(1)
        self.lock_img('img/zhucaidan/bangzhu.bmp', elseclick=[(32, 32)])  # 锁定帮助
        pcr_log(self.account).write_log(level='info', message='账号：%s已修改名字' % name)

    def get_bar(self, bar: PCRelement, screen=None):
        """
        进度条类百分比获取
        :param bar: 含有at,fc,bc元素的PCRelement
            其中,at为截取进度条，fc为进度条【横向中线】前景色，bc为进度条【横向中线】背景色
        :param screen: 设置为None，重新截屏
        :return: 百分比0~1
        """
        if screen is None:
            screen = self.getscreen()
        at, fc, bc = bar.at, bar.fc, bar.bc
        x1, y1, x2, y2 = at
        ym = int((y1 + y2) / 2)  # 只取中之条
        mid_line = UIMatcher.img_cut(screen, (x1, ym, x2, ym))
        # R,G,B -> B G R
        fc = np.array([fc[2], fc[1], fc[0]])
        bc = np.array([bc[2], bc[1], bc[0]])
        tf = np.sqrt(((mid_line - fc) ** 2).sum(axis=2)).ravel()
        tb = np.sqrt(((mid_line - bc) ** 2).sum(axis=2)).ravel()
        t = tf < tb
        left = 0
        right = len(t) - 1
        for ind in range(len(t)):
            if t[ind]:
                left = ind
                break
        for ind in range(len(t) - 1, -1, -1):
            if not t[ind]:
                right = ind
                break
        t = t[left:right + 1]
        return t.sum() / len(t)

    def get_daoju_number(self, screen=None, must_int=True, do_addition_check=True):
        """想尽一切办法获得右上角道具数量。
        利用x号定位，获取精确范围。
        若开启must_int：则会再搞不出整数时返回(None, 原始str），搞出时返回（整数，原始str）
        否则，返回整数或原始str
        """
        MIDS = {1: 49, 2: 39, 3: 31, 4: 21, 5: 11}  # 中位数

        sc = self.getscreen() if screen is None else screen
        at = (647, 199, 714, 217)
        sc = UIMatcher.img_cut(sc, at=at)
        plus = cv2.imread(filename="img/plus.bmp")
        choose = self.img_where_all_prob(plus, screen=sc, threshold=0.6)
        if len(choose) == 0:
            if must_int:
                return None, -1
            else:
                return -1
        choose = choose[0]
        prob, x, y, (x1, y1, x2, y2) = choose
        num_at = (x2 + 647 + 4, 199, 720, 214)
        out = self.ocr_int(*num_at)
        if use_pcrocr_to_process_basic_text:
            if must_int:
                return out, str(out)  # 这个准啊！
            else:
                return str(out)
        if out == -1:
            if must_int:
                return None, out
            else:
                return out
        new_out = make_it_as_number_as_possible(out)
        if len(new_out) == 0:
            if must_int:
                return None, out
            else:
                return out
        the_int = int(new_out)
        int_len = len(str(the_int))
        if int_len > 5:
            if must_int:
                return None, out
            else:
                return the_int
        # The median X should be +- 3
        M_X = MIDS[int_len]
        if -3 <= x - M_X <= 3:
            # Good int, maybe.
            if must_int:
                return the_int, out
            else:
                return the_int
        else:
            # Maybe Bad INT.
            if must_int:
                return None, out
            else:
                return must_int

    def kucunshibie(self, scan_zb=True, scan_sp=True, var: Optional[dict] = None):
        self.check_ocr_running()
        mv = movevar(var)
        self.lock_home()
        title_at = (613, 85, 909, 112)
        self.lock_img(ZHUCAIDAN_BTN["bangzhu"], elseclick=[(871, 513)])  # 锁定帮助
        # 去道具
        self.lock_no_img(ZHUCAIDAN_BTN["daoju"], elseclick=[(536, 159)])
        self.lock_img(ZHUCAIDAN_BTN["daojuyilan"], elseclick=[(536, 159)])  # 锁定道具一览

        LAST_PAGE = False

        def get_equ_at(r, c):
            EQU_X = [97, 203, 315, 421, 535]
            EQU_Y = [126, 228, 336]
            if LAST_PAGE:
                EQU_Y = [198, 305, 412]
            return EQU_X[c], EQU_Y[r]

        DIR = ""
        LAST_SCREEN = None

        def dao_ju_kuang(screen=None):
            at = (616, 78, 924, 227)  # 道具框
            djk = screen if screen is not None else self.getscreen()
            djk = UIMatcher.img_cut(djk, at)
            return djk

        def check_last_screen():
            # 防止同一屏幕重复出现
            nonlocal LAST_SCREEN
            if LAST_SCREEN is None:
                LAST_SCREEN = dao_ju_kuang(self.last_screen)
                return True
            else:
                NOW_SCREEN = dao_ju_kuang(self.last_screen)
                if self.img_equal(NOW_SCREEN, LAST_SCREEN) > 0.98:
                    return False
                else:
                    LAST_SCREEN = NOW_SCREEN
                    return True

        def output_dict(d):
            path = os.path.join("outputs", DIR)
            if not os.path.isdir(path):
                os.makedirs(path)
            sd = sorted(d)
            with open(os.path.join(path, self.account + ".txt"), "w", encoding="utf-8") as f:
                f.write("%s\t%s\t%s\t%s\n" % ("名称", "数量", "更新时间", "备注"))
                for k in sd:
                    f.write("%s\t%s\t%s\t%s\n" % (k, d[k][0], get_time_str(d[k][1]), d[k][2]))

        def output_warning_pic(title, value):
            path = os.path.join("outputs", DIR, "warning", self.account)
            if not os.path.isdir(path):
                os.makedirs(path)
            target = os.path.join(path, title + ".bmp")
            djk = dao_ju_kuang()
            cv2.imwrite(target, djk)
            self.log.write_log("warning", f"在识别{title}时读到了不可识别的{value}，图片已保存至{target}")

        def getrecord():
            kucun = self.AR.get(DIR, {})
            return kucun

        def addrecord(d, nam, val, bz=""):
            d[nam] = (val, time.time(), bz)

        def saverecord(d):
            self.AR.set(DIR, d)

        def get_number_by_sale():
            sc = self.getscreen()
            if self.is_exists(ZHUCAIDAN_BTN["sale_short"], screen=sc):
                self.click_btn(ZHUCAIDAN_BTN["sale_short"], until_appear=ZHUCAIDAN_BTN["chushouqueren"])
            elif self.is_exists(ZHUCAIDAN_BTN["sale_long"], screen=sc):
                self.click(ZHUCAIDAN_BTN["sale_long"], until_appear=ZHUCAIDAN_BTN["chushouqueren"])
            else:
                return None
            sc = self.last_screen
            for _ in range(6):
                self.click(1, 1)
            at = (492, 266, 566, 286)
            out = self.ocr_int(*at, screen_shot=sc)
            new_out = make_it_as_number_as_possible(out)
            try:
                the_int = int(new_out)
                return the_int
            except:
                return None

        def dragdown():
            obj = self.d.touch.down(55, 445)
            time.sleep(0.5)
            obj.move(55, 130)
            time.sleep(0.8)
            sc = self.getscreen()
            r1c0 = UIMatcher.img_cut(sc, at=(56, 354, 140, 441))
            r1c0.std()
            flag = False
            if r1c0.std() < 15:
                # 拖到底了
                flag = True
            obj.up(55, 130)
            time.sleep(1)
            return flag

        if scan_zb and mv.notflag("zb_scanned"):
            # 扫描装备
            DIR = "zhuangbei_kucun"
            rec = getrecord()
            self.lock_img(ZHUCAIDAN_BTN["sortico"], elseclick=ZHUCAIDAN_BTN["zhuangbei"])
            mv.regflag("zb_r", 0)  # 行数
            mv.regflag("zb_c", 0)  # 列数
            mv.regflag("zb_p", 0)  # 页数
            LAST_PAGE = mv.flag("zb_last_page")
            for _ in range(var["zb_p"]):
                dragdown()  # 回到上次页数
            while True:
                while var["zb_r"] < 3:
                    count = 0
                    while var["zb_c"] < 5:
                        if count >= 25 or (count >= 5 and not fast_screencut) or (
                                count >= 10 and mv.flag("zb_last_page")):
                            self.log.write_log("warning", "不反映了，可能结束了。")
                            var["zb_c"] = 999
                            var["zb_r"] = 999
                            break
                        x, y = get_equ_at(var["zb_r"], var["zb_c"])
                        self.click(x, y, post_delay=0.5 * (count == 0) + 0.1 + 5 * (count % 10 == 9))
                        sc = self.getscreen()
                        if not check_last_screen():
                            count += 1
                            continue
                        title = self.ocr_center(*title_at, screen_shot=sc)
                        if title == -1:
                            count += 1
                            continue
                        title = make_it_as_zhuangbei_as_possible(title)
                        title = self._check_img_in_list_or_dir(title, (616, 76, 884, 194), "ocrfix/zb", "EQU_ID", sc)
                        out, original_out = self.get_daoju_number(sc, True)
                        comment = ""
                        if out is None:
                            out = get_number_by_sale()
                        if out is None:
                            # 没救了
                            out = original_out
                            output_warning_pic(title, out)
                            comment = "存疑"
                        addrecord(rec, title, out, comment)
                        saverecord(rec)
                        var["zb_c"] += 1
                        mv.save()
                    if var["zb_c"] == 999:
                        break
                    var["zb_c"] = 0
                    var["zb_r"] += 1
                    mv.save()
                if var["zb_r"] == 999:
                    mv.setflag("zb_scanned")
                    break
                flag = dragdown()

                if flag:
                    if mv.notflag("zb_last_page"):
                        mv.setflag("zb_last_page")
                        LAST_PAGE = True
                    else:
                        mv.setflag("zb_scanned")
                        break
                time.sleep(1)
                LAST_SCREEN = dao_ju_kuang()
                var["zb_r"] = 0
                var["zb_p"] += 1
                mv.save()
            # Output
            output_dict(rec)
        mv.clearflags()
        self.lock_home()

    def count_stars(self, star_dict=None, screen=None):
        """
        获取右上角当前关卡的星星数
        :param screen: 设置为None时，不另外截屏
        :return: 0~3
        """
        if screen is None:
            screen = self.getscreen()
        fc = np.array([98, 228, 245])  # G B R:金色
        bc = np.array([212, 171, 139])  # G B R:灰色
        c = []
        us = star_dict
        for i in range(min(star_dict), max(star_dict) + 1):
            x = us[i].x
            y = us[i].y
            c += [screen[y, x]]
        c = np.array(c)
        tf = np.sqrt(((c - fc) ** 2)).sum(axis=1)
        tb = np.sqrt(((c - bc) ** 2)).sum(axis=1)
        t = tf < tb
        return np.sum(t)

    def _load_data_cache(self):
        if hasattr(self, "data_cache"):
            data = getattr(self, "data_cache")
        else:
            from DataCenter import LoadPCRData
            data = LoadPCRData()
            if data is not None:
                setattr(self, "data_cache", data)
        return data

    def _check_img_in_list_or_dir(self, target_txt, target_pic_at, target_dir, target_list_name, screen):

        data = self._load_data_cache()
        if data is None:
            return target_txt  # No Dataset, Do Nothing.
        target_list = getattr(data, target_list_name)
        if target_txt in target_list:
            return target_txt  # Good
        if not os.path.isdir(target_dir):
            os.makedirs(target_dir)
        P = pathlib.Path(target_dir)
        if target_pic_at is not None:
            screen = UIMatcher.img_cut(screen, target_pic_at)
        for p in P.iterdir():
            if p.suffix == ".bmp":
                bmp2 = cv2.imdecode(np.fromfile(str(p), dtype=np.uint8), -1)
                if self.img_equal(screen, bmp2, similarity=0.1) > 0.98:
                    if debug:
                        self.log.write_log('debug', f"找到相似图片：{p}")
                    if p.stem in target_list:
                        return p.stem

        # 失败
        target_name = checkNameValid(target_txt)
        save_target = os.path.join(target_dir, target_name + ".bmp")
        save_target = str(pathlib.Path(save_target))
        cv2.imencode('.bmp', screen)[1].tofile(save_target)
        self.log.write_log("warning", f"文字{target_txt}可能识别有误！请修改{save_target}的文件名为正确的值！")
        return target_txt

    def jueseshibie(self, var: Optional[dict] = None):
        mv = movevar(var)
        self.check_ocr_running()
        S = self.get_zhuye().goto_juese()
        self.click(299, 23)  # 全部
        S = S.enter_first_juese()

        def output_dict(d):
            path = os.path.join("outputs", "juese_info")
            if not os.path.isdir(path):
                os.makedirs(path)
            sd = sorted(d)
            with open(os.path.join(path, self.account + ".txt"), "w", encoding="utf-8") as f:
                f.write("\t".join(["名称", "星级", "Rank", "等级", "左上", "右上", "左中", "右中", "左下", "右下", "好感", "更新时间"]) + "\n")
                for k in sd:
                    v = d[k]
                    f.write("\t".join([str(s) for s in [k, v["star"], v["rank"], v["dengji"], *v["zb"], v["haogan"],
                                                        get_time_str(v["last_update"])]]) + "\n")

        mv.regflag("count", 0)
        S = S.goto_kaihua()
        FIRST_NAME = S.get_name()
        S = S.goto_zhuangbei()

        for _ in range(var["count"]):
            S.next_char()

        while True:
            sc = self.getscreen()
            data = self.AR.get("juese_info", UDD["juese_info"])
            # Main Info
            D = {}
            D["haogan"] = S.get_haogan(sc)
            D["dengji"] = S.get_level(sc)
            D["rank"] = S.get_rank(sc)
            D["zb"] = S.get_six_clothes(sc)
            S = S.goto_kaihua()
            sc = self.getscreen()
            NAME = S.get_name(sc)
            if NAME == FIRST_NAME and var["count"] != 0:
                break
            D["star"] = S.get_stars(sc)
            D["last_update"] = time.time()
            if NAME not in data:
                data[NAME] = {}
            data[NAME].update(D)
            self.AR.set("juese_info", data)
            var["count"] += 1
            mv.save()
            S = S.goto_zhuangbei()
            S.next_char()
        mv.clearflags()
        output_dict(self.AR.get("juese_info", UDD["juese_info"]))
        self.lock_home()

    def guojuqing(self, story_type=""):
        while True:
            screen = self.getscreen()
            lst = self.img_where_all(img="img/juqing/xuanzezhi_1.bmp", at=(233, 98, 285, 319), screen=screen)
            self.log.write_log('info ', f"{lst}")
            # 选择无语音选项
            if self.is_exists(JUQING_BTN["wuyuyin"].img, screen=screen, at=(410, 277, 553, 452)):
                self.click_img(img=JUQING_BTN["wuyuyin"].img, screen=screen, at=(410, 277, 553, 452))
                continue
            # 选择快进剧情
            if self.is_exists(JUQING_BTN["caidanyuan"], screen=screen):
                self.click_btn(JUQING_BTN["caidanyuan"], until_appear=(JUQING_BTN["tiaoguo_1"]))
                # 快进确认弹出
                self.click_btn(JUQING_BTN["tiaoguo_1"], until_appear=(JUQING_BTN["tiaoguo_2"]))
                continue
            # 确认快进，包括视频和剧情
            if self.is_exists(JUQING_BTN["tiaoguo_2"], screen=screen):
                self.click_btn(JUQING_BTN["tiaoguo_2"])
                continue
            # 选择支固定选红色
            if len(lst) > 0:
                self.click(int(lst[0]), int(lst[1]))
                continue

            # 三种退出形式
            # 报酬确认 (好感度剧情)
            if self.is_exists(JUQING_BTN["baochouqueren"], at=(433, 73, 523, 100),
                              screen=screen) and story_type == "haogandu":
                '''
                433, 73, 523, 100
                433, 28, 523, 55
                '''
                self.click(475, 429)  # 点击关闭
                self.log.write_log('info', "完成了这段剧情")
                break
            if self.is_exists(JUQING_BTN["baochouqueren"], at=(433, 28, 523, 55),
                              screen=screen) and story_type == "haogandu":
                '''
                433, 73, 523, 100
                433, 28, 523, 55
                '''
                self.click(475, 473)  # 点击关闭
                self.log.write_log('info', "完成了这段剧情")
                break

            # 主线剧情退出检测
            if self.is_exists(JUQING_BTN["guanbi"], screen=screen) and story_type == "zhuxian":
                self.click_btn(JUQING_BTN["guanbi"])
                time.sleep(1)
                self.fclick(1, 1)
                self.log.write_log('info', "完成了这段剧情")
                break
            # 兼容信赖度退出检测，不是很稳定，因为点无语音的时候背景也有
            if story_type == "xianlai" and self.is_exists("img/juqing/new_content.bmp", screen=screen, threshold=0.7):
                self.log.write_log('info', "完成了这段剧情")
                break
            else:
                self.fclick(479, 260)

    def check_color(self, fc, bc, xcor, ycor, color_type="gbr", screen=None):
        # 主要用于检测点的颜色是否为前景色，通过比较RGB值与前景色/背景色的距离
        # fc:前景色 bc:背景色 xcor:点坐标x ycor:点坐标y color_type:颜色格式
        if screen is None:
            screen = self.getscreen()
        if color_type == "rgb":
            fc[0], fc[1], fc[2] = fc[2], fc[0], fc[1]
            bc[0], bc[1], bc[2] = bc[2], bc[0], bc[1]
        fc = np.array(fc)
        bc = np.array(bc)
        c = screen[ycor, xcor]
        tf = ((c - fc) ** 2).sum()
        tb = ((c - bc) ** 2).sum()
        if tf < tb:
            return True
        else:
            return False
